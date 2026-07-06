import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook("D:/Zalo/TCC_Template Request Form_2026 version.xlsm", data_only=True)

ws = wb['Requestor View']
print("=" * 60)
print("SHEET: Requestor View")
print("=" * 60)

for r in range(1, 35):
    row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
    if any(v is not None for v in row_vals):
        clean_row = [str(x).replace('\n', ' ') if x is not None else "" for x in row_vals]
        print(f"Row {r:02d}: {clean_row}")
